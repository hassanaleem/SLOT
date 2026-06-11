#!/usr/bin/env python3
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side


BASE = Path(__file__).resolve().parent
CSV_PATH = BASE / "time_totals.csv"
XLSX_PATH = BASE / "time_totals_highlighted.xlsx"


def to_float(value):
    try:
        return float(value)
    except ValueError:
        return None


with CSV_PATH.open(newline="") as f:
    rows = list(csv.DictReader(f))

headers = ["benchmark"] + [h for h in rows[0].keys() if h.endswith("_seconds")]

wb = Workbook()
ws = wb.active
ws.title = "time_totals"

fastest_fill = PatternFill(fill_type="solid", fgColor="FFFFF2CC")
thin_side = Side(style="thin", color="FFB7B7B7")
row_side = Side(style="medium", color="FF808080")
row_end_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=row_side)

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.border = row_end_border

for row_num, row in enumerate(rows, start=2):
    values = {h: to_float(row[h]) for h in headers if h.endswith("_seconds")}
    fastest = min(v for v in values.values() if v is not None)

    for col, header in enumerate(headers, start=1):
        value = values.get(header) if header in values else row[header]
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.border = row_end_border

        if header in values:
            cell.number_format = "0.0000"
            if value == fastest:
                cell.fill = fastest_fill
                cell.font = Font(bold=True)

ws.freeze_panes = "B2"
ws.auto_filter.ref = ws.dimensions

for column_cells in ws.columns:
    width = max(len(str(cell.value)) for cell in column_cells) + 2
    ws.column_dimensions[column_cells[0].column_letter].width = min(width, 42)

wb.save(XLSX_PATH)
print(f"Wrote {XLSX_PATH}")
